#!/usr/bin/python3
#-*- coding=utf8 -*-
#文件说明：
import openpyxl

class OperExcel:
    #生成文件对象和工作薄对象
    def __init__(self,filename,sheetname):
        '''
        :param filename: Excel的文件名
        :param sheetname: Excel的工作薄名
        '''
        self.filename = filename
        self.wb = openpyxl.load_workbook(self.filename)
        self.sheet = self.wb[sheetname]

    def __del__(self):
        self.wb.close()
    #获取最大行数
    def maxrow(self):
        maxrow = self.sheet.max_row
        return maxrow
    def maxcol(self):
        maxcolnum = self.sheet.max_column
        return maxcolnum

    #在指定单元格写入一个数据
    def wtdata(self,rownum,colnum,msg):
        self.sheet.cell(row=rownum,column=colnum).value = msg
        self.wb.save(self.filename)

    #在末尾行追加一行数据,data为一个列表
    def wtline(self,*data):
        maxrow = self.sheet.max_row
        for cn in range(len(data)):
            self.sheet.cell(row=maxrow+1,column=cn+1).value = data[cn]
        self.wb.save(self.filename)

    #获取某行的值，无参时获取标题
    def rdrow(self,rownum=1):
        max_colnum = self.sheet.max_column
        row_data = []
        for i in range(max_colnum):
            value = self.sheet.cell(row=rownum,column=i+1).value
            row_data.append(value)
        return row_data

    #获取一整列数据的值
    def rdcolumn(self,colnum=1):
        max_row = self.sheet.max_row
        colnum_data = []
        for i in range(max_row):
            value = self.sheet.cell(row=i+1,column=colnum).value
            colnum_data.append(value)
        return colnum_data

    #读取整个sheet页中的数据
    def rdall(self):
        '''
        :通过self.rdrow()方法获取每行数据
        :然后再依次追加到总表alldata中来即可
        '''
        max_row = self.sheet.max_row
        alldata = []
        for rownum in range(max_row):
            alldata.append(self.rdrow(rownum+1))
        return alldata